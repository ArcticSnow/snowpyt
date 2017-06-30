# script to develop code for reading directly from the xlsx file

from __future__ import division
import numpy as np
import pandas as pd
from openpyxl import load_workbook

file = '/snowpyt/data_example/20170209_Finse_snowpit.xlsx'

# 1. try using pandas (intall xlrd)
xl = pd.read_excel(file, sheetname=, header=, skiprows=)



import openpyxl as xl

a = xl.load_workbook('data_example/20170209_Finse_snowpit.xlsx')
s = a.get_sheet_by_name('Sheet1')



def csv_from_excel():
    wb = xlrd.open_workbook('your_workbook.xls')
    sh = wb.sheet_by_name('Sheet1')
    your_csv_file = open('your_csv_file.csv', 'wb')
    wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

    for rownum in xrange(sh.nrows):
        wr.writerow(sh.row_values(rownum))

    your_csv_file.close()


import xlrd


def find_row_with_values(sheet, val):
    '''
    Function to find cell location of particular values indicated by a field
    :param sheet: excel sheet (open with xlrd library)
    :param val: values of the fields (the first 4 character for each field) in a list format
    :return: return a dictionnary of the row where the value wanted is
    '''

    rows = []
    values = []
    for row_index in xrange(sheet.nrows):
        for value in val:
            if value == (str(sheet.row(row_index)[0].value)[:4]):
                rows.append(row_index)
                values.append(value)
    print val
    print rows
    return dict(zip(values, rows))


def get_cell_val_str(sh, cRrow, cCol):
    '''
    Function to grab cell value as str
    :param sh:
    :param cRrow:
    :param cCol:
    :return:
    '''
    value = sh.cell(cRrow, cCol)
    if value.value.__len__() == 0:
        value.value = np.nan
    return str(value.value)

file = 'snowpyt/data_example/20170209_Finse_snowpit.xlsx'
wb = xlrd.open_workbook(file)
sh = wb.sheet_by_name('Sheet1')

fields = ['East', 'Nort','Elev', 'Date', 'Obse', 'Loca', 'Air ', 'Weat', 'Comm', 'Snow', 'Time', 'Gene', 'Stra']
values = find_row_with_values(sh, fields)

get_cell_val_str(sh, 0, 1)
import pandas as pd
profile_raw_table = pd.read_excel(file,sheet=sh, skiprows=int(values.get('Stra'))+1, engine='xlrd')

plots_order=['temperature', 'stratigraphy', 'hardness', 'crystal size', 'density']
plots_dict = {'temperature':1,
                            'density':2,
                            'stratigraphy':3,
                            'hardness':4,
                            'crystal size':5}





















# 2. try unsing openpyxl
xlsx = load_workbook(file)





'''
Example for loading the csv file:

    def load_profile(self):
        f = open(self.filename)
        for k,line in enumerate(f):
            if line[0:12] == 'Stratigraphy':
                break
        f.close()

        self.profile_raw_table = pd.read_csv(self.filename, sep='\t', skiprows=k+1)
        self.layerID = self.profile_raw_table['Layer ID']
        self.layer_top = self.profile_raw_table['Top [cm]']
        self.layer_bot = self.profile_raw_table['Bottom [cm]']

        self.grain_type1 = self.profile_raw_table['Type 1']
        self.grain_type2 = self.profile_raw_table['Type 2']
        self.grain_type3 = self.profile_raw_table['Type 3']
        self.grain_size_min = self.profile_raw_table['Diameter min [mm]']
        self.grain_size_max = self.profile_raw_table['Diameter max [mm]']

        self.hardness = self.profile_raw_table['Hardness']
        self.hardness_code = self.profile_raw_table['Hardness code']

        self.density_depth = self.profile_raw_table['Depth Center [cm]']
        self.density = self.profile_raw_table['Snow Density [g/cm3]']

        self.temperature_depth = self.profile_raw_table['Depth [cm]']
        self.temperature_snow = self.profile_raw_table['Temp [deg C]']

        self.sample_depth = self.profile_raw_table['Depth Center [cm].1']
        self.sample_name = self.profile_raw_table['ID_sample'].astype(str)


    def load_metadata(self):
        f = open(self.filename)
        try:
            for i, line in enumerate(f):
                if line[0:4] == 'Date':
                    self.date = line.split("\t")[1]
                if line[0:4] == 'Time':
                    self.Time = line.split("\t")[1]
                if line[0:4] == 'Gene':
                    self.General_loc = line.split("\t")[1]
                if line[0:4] == 'East':
                    self.East = line.split("\t")[1]
                    self.East_unit = line.split("\t")[2]
                if line[0:4] == 'Nort':
                    self.North = line.split("\t")[1]
                    self.North_unit = line.split("\t")[2]
                if line[0:4] == 'Elev':
                    self.Elevation = line.split("\t")[1]
                    self.Elevation_unit = line.split("\t")[2]
                if line[0:4] == 'Obse':
                    self.Observer = line.split("\t")[1]
                if line[0:4] == 'Air ':
                    self.AirTemp = line.split("\t")[1]
                if line[0:4] == 'Weat':
                    self.weather_conditions = line.split("\t")[1]
                if line[0:4] == 'Comm':
                    self.comments = line.split("\t")[1]
        except ValueError:
            print "Could not load metadata. Check file formating"
        f.close()